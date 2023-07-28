import pandas as pd
from nltk.sentiment.vader import SentimentIntensityAnalyzer
from openpyxl import load_workbook

# dates = ['08-28', '08-29', '08-30', '08-31', '09-01', '09-02', '09-03', '09-04']
dates = ['08-28']
analyzer = SentimentIntensityAnalyzer()
all_tweets = []
for date in dates:
    wb = load_workbook('/Users/ericli/Downloads/Hurricane Ida/storm ida_2021-' + date + '.xlsx')
    ws = wb.active
    for row in range(2, ws.max_row):
        all_tweets.append(ws['I' + str(row)].value)
    df = pd.DataFrame({'tweets': all_tweets})
    df['neg'] = df['tweets'].apply(lambda x: analyzer.polarity_scores(x)['neg'])
    df['neu'] = df['tweets'].apply(lambda x: analyzer.polarity_scores(x)['neu'])
    df['pos'] = df['tweets'].apply(lambda x: analyzer.polarity_scores(x)['pos'])
    df['compound'] = df['tweets'].apply(lambda x: analyzer.polarity_scores(x)['compound'])
    print(df)
    print(df['compound'].describe())
