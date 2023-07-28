from wordcloud import WordCloud, STOPWORDS
import matplotlib.pyplot as plt
from openpyxl import load_workbook

dates = ['08-28', '08-29', '08-30', '08-31', '09-01', '09-02', '09-03', '09-04']

# for date in dates:
wb = load_workbook('/Users/ericli/Downloads/Hurricane Ida/storm ida_2021-' + dates[0] + '.xlsx')
ws = wb.active
text = ""
for row in range(2, ws.max_row):
    text = text + ws['I' + str(row)].value.lower()

stopwords = set(STOPWORDS)
stopwords.update(["https", "t", "co"])
wordcloud = WordCloud(width=800, height=600, stopwords=stopwords, background_color="white").generate(text)
plt.figure()
plt.imshow(wordcloud, interpolation='bilinear')
plt.axis("off")
plt.show()
# plt.savefig('ida_' + date + '_cloud.png')
