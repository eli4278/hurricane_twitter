from openpyxl import load_workbook
import matplotlib.pyplot as plt
from scipy import stats

# dates = ['8/26', '8/27', '8/28', '8/29', '8/30', '8/31', '9/1', '9/2', '9/3', '9/4', '9/5', '9/6', '9/7', '9/8', '9/9', '9/10', '9/11', '9/12', '9/13']
# number_of_tweets = [5603, 29362, 27086, 86797, 51808, 14014, 7520, 36031, 16220, 6531, 4756, 3557, 4688, 3069, 1460, 1140, 533, 1269, 1267]
# print(stats.describe(number_of_tweets))
# plt.xticks(range(len(number_of_tweets)), dates)
# plt.xlabel('Date')
# plt.ylabel('Number of Tweets')
# plt.title('Storm Ida Tweets per Day 8/26 to 9/13')
# plt.bar(range(len(number_of_tweets)), number_of_tweets)
# plt.show()
dates = []
number_of_tweets = []
for x in range(21):
    month = ''
    day = ''
    if x > 7:
        month = '9'
        if x < 17:
            day = '0' + str(x - 7)
        else:
            day = str(x - 7)
    else:
        month = '8'
        day = str(x + 24)
    filename = 'Hurricane Ida/hurricane_2021-0' + month + '-' + day + '.xlsx'
    wb = load_workbook(filename)
    ws = wb.active
    max_rows = ws.max_row
    print(month + '-' + day + ': ' + str(max_rows) + ' tweets')
    dates.append(month + '/' + day)
    number_of_tweets.append(max_rows)

print(stats.describe(number_of_tweets))
plt.xticks(range(len(number_of_tweets)), dates)
plt.xlabel('Date')
plt.ylabel('Number of Tweets')
plt.title('Hurricane Tweets per Day 8/24 to 9/13')
plt.bar(range(len(number_of_tweets)), number_of_tweets)
plt.show()
