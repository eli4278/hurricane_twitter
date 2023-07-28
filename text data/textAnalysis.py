# from nltk.tokenize import word_tokenize
# from nltk.probability import FreqDist
import matplotlib.pyplot as plt
from openpyxl import load_workbook

dates = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13']
august_dates = ['26', '27', '28', '29', '30', '31']
tweets = []
for day in dates:
    print('############################################')
    print('DAY: ' + 'SEPTEMBER ' + day)
    print()
    wb = load_workbook('Hurricane Ida/storm ida_2021-09-' + day + '.xlsx')
    ws = wb.active
    keywords = ['stress', 'beds', 'flooding', 'outage', 'shutdown']

    for word in keywords:  # find keywords in tweets
        tweets = []
        print('KEYWORD: ' + word)
        print()
        count = 0
        for row in range(2, ws.max_row):
            tweet = ws['I' + str(row)].value
            time = ws['G' + str(row)].value
            if word in tweet.lower() and tweet not in tweets:
                count = count + 1
                print('Tweet #: ' + str(row - 2) + ' Time: ' + time)
                print('Username: ' + ws['B' + str(row)].value + ' Location: ' + str(ws['E' + str(row)].value))
                print(tweet)
                print()
                tweets.append(tweet)
        print('Count: ' + str(count))
        print()

# count retweets for each day
# for x in range(19):
#     month = ''
#     day = ''
#     if x > 5:
#         month = '9'
#         if x < 15:
#             day = '0' + str(x - 5)
#         else:
#             day = str(x - 5)
#     else:
#         month = '8'
#         day = str(x + 26)
#     filename = 'Hurricane Ida/storm ida_2021-0' + month + '-' + day + '.xlsx'
#     wb = load_workbook(filename)
#     ws = wb.active
#     print(month + '-' + day)
#     count = 0
#     for row in range(2, ws.max_row):
#         if ws['I' + str(row)].value[0:2] == 'RT':
#             count = count + 1
#     print('tweets: ' + str(ws.max_row - 2) + ' retweets: ' + str(count))
#     print('percentage of retweets: ' + str(count / (ws.max_row - 2) * 100))


# text = """Hello Mr. Smith, how are you doing today? The weather is great, and city is awesome.
# The sky is pinkish-blue. You shouldn't eat cardboard"""
# list = ['smith', 'blue']
# for word in list:
#     if word in text.lower():
#         print(True)
# tokenized_word = word_tokenize(text)
# print(tokenized_word)
# fdist = FreqDist(tokenized_word)
# print(fdist)
# fdist.plot(30, cumulative=False)
# plt.show()
